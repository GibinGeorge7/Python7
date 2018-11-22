
"""""
#read all rows of columns B  in sheet1

import openpyxl
workbook=openpyxl.load_workbook("example.xlsx")
sheet=workbook['Sheet1']
for i in  range(1,10):
    if str(sheet.cell(row=i, column=2).value) != 'None':
        print (sheet.cell(row=i, column=2).value)
"""""

# CREATE EXCEL WorkBook
import openpyxl
wb=openpyxl.Workbook()
sheet1=wb[wb.sheetnames[0]]
sheet1['A1']=435678
sheet1['A2']="Hello"
#create new sheet "Data
wb.create_sheet(title='Data')
sheet2=wb[wb.sheetnames[1]]
sheet2['B1']=99999
sheet2['B2']="Hello1"
#overwites existing sheet or create NEW
wb.save("example1.xlsx")



