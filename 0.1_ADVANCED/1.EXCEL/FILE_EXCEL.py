import openpyxl

workbook=openpyxl.load_workbook("example.xlsx")

#find all sheets in workbook
print(workbook.sheetnames)



from openpyxl import load_workbook
wb = load_workbook(filename = 'example.xlsx')
ws=wb.active
print(ws['A1'].value)
