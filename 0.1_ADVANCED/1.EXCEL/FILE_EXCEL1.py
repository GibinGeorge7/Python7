import openpyxl

#load workbook
workbook=openpyxl.load_workbook("example.xlsx")
#find all sheets in workbook
print(workbook.sheetnames)
#access Sheet1
sheet=workbook['Sheet1']


#Read CELL B1 of sheet1
print(sheet['B1'].value)

#read sheet CELL B1 of sheet1
print(sheet.cell(row=1,column=2).value)


